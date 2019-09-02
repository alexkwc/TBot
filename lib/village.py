import urllib

import openpyxl


class Village:
    def __init__(self, driver):
        self.driver = driver

    async def get_resource_info(self):
        return self.driver.execute_script("return resources;")

    async def getIds(self):
        hrefs = self.driver.find_elements_by_xpath(
            f'//a[starts-with(@href, "?newdid=")]'
        )
        villageIds = []
        for ref in hrefs:
            href = ref.get_attribute("href")
            villageIds.append(
                urllib.parse.parse_qs(
                    urllib.parse.urlparse(href).query
                )['newdid'][0]
            )
        return villageIds

    async def createScheduler(self, villageIds):
        print(villageIds)
        if len(villageIds) < 1:
            return None

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = villageIds[0]
        for index in range(1, len(villageIds)):
            workbook.create_sheet(villageIds[index])
        return workbook

    async def setScheduler(self, worksheet, fieldInfo):
        col = 2
        row = 2
        for fieldName, fieldLevels in fieldInfo.items():
            worksheet.cell(row=row, column=col, value=fieldName)
            worksheet.cell(row=row + 1, column=col, value='Current Level')
            worksheet.cell(row=row + 2, column=col, value='Target Level')
            row += 1
            for index in range(0, len(fieldLevels)):
                worksheet.cell(
                    row=row, column=col + index + 1, value=fieldLevels[index]
                )
            row += 3
        return True

    async def getScheduler(self, worksheet, fieldInfo):
        todo = []
        col = 2
        row = 2
        for fieldName, fieldLevels in fieldInfo.items():
            row += 1
            for index in range(0, len(fieldLevels)):
                currentLevel = fieldLevels[index]
                targetLevel = worksheet.cell(
                    row=row + 1, column=col + index + 1
                ).value
                if not targetLevel:
                    continue
                if targetLevel > currentLevel:
                    todo.append((fieldName, currentLevel))

            row += 3
        return todo
